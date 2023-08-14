import openpyxl
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime
import time

# Excel file read
excel = 'data.xlsx'

# List of sheet names
sheet_names = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
current_day = datetime.today().weekday()
current_day_sheet = sheet_names[current_day]
workbook = openpyxl.load_workbook(excel)
sheet = workbook[current_day_sheet]

driver = webdriver.Chrome()


driver.get('https://www.google.com')


try:
    search_box = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.NAME, 'q')))
except TimeoutException:
    print("Search box not found on the page.")
    driver.quit()
    exit()

# Collect data for the suggestions
suggestions_data = []

for row_idx in range(3, 14):
    # Get the value from column 3 (index 2)
    search_query = sheet.cell(row=row_idx, column=3).value

    #search query
    if search_query:
        search_box.clear()
        search_box.send_keys(search_query)
        time.sleep(5) #search list wait

        # inspect path
        try:
            suggestions_box = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/div[3]/form/div[1]/div[1]/div[2]/div["
                                                          "2]/div[2]/div[1]/div/ul"))
            )
            suggestions_elements = suggestions_box.find_elements(By.CSS_SELECTOR, 'li.sbct')
            suggestions = [element.text for element in suggestions_elements]

            # Find long and short
            longest_suggestion = max(suggestions, key=len, default=search_query)
            shortest_suggestion = min(suggestions, key=len, default=search_query)

            # store in excel
            sheet.cell(row=row_idx, column=4, value=longest_suggestion)
            sheet.cell(row=row_idx, column=5, value=shortest_suggestion)

        except TimeoutException:
            print("not found:", search_query)

# Save the changes to the Excel file
workbook.save(excel)

# Close the browser
driver.quit()
